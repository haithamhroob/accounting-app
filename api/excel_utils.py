"""
Excel Utility Functions
دوال مساعدة لـ Excel
"""
import io
from datetime import datetime
from decimal import Decimal

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def check_excel_dependencies():
    """Check if Excel dependencies are installed"""
    if not EXCEL_AVAILABLE:
        raise ImportError(
            "Excel dependencies not installed. Run: pip install openpyxl pandas"
        )


def style_header_row(ws, row_num=1):
    """Apply header styling to a row"""
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for cell in ws[row_num]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment


def auto_adjust_columns(ws):
    """Auto-adjust column widths based on content"""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def create_parties_excel(parties_queryset):
    """Create Excel file for parties"""
    check_excel_dependencies()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "الأطراف"
    ws.sheet_view.rightToLeft = True  # RTL for Arabic
    
    # Headers
    headers = ['المعرف', 'الاسم', 'النوع', 'الهاتف', 'البريد الإلكتروني', 'الرصيد', 'تاريخ الإضافة']
    ws.append(headers)
    style_header_row(ws)
    
    # Data
    for party in parties_queryset:
        ws.append([
            party.id,
            party.name,
            party.get_party_type_display(),
            party.phone or '',
            party.email or '',
            str(party.get_balance()),
            party.created_at.strftime('%Y-%m-%d %H:%M') if party.created_at else ''
        ])
    
    auto_adjust_columns(ws)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_invoices_excel(invoices_queryset):
    """Create Excel file for invoices"""
    check_excel_dependencies()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "الفواتير"
    ws.sheet_view.rightToLeft = True
    
    # Headers
    headers = ['رقم الفاتورة', 'الطرف', 'النوع', 'تاريخ الإصدار', 'تاريخ الاستحقاق', 
               'الحالة', 'المجموع', 'المدفوع', 'المتبقي']
    ws.append(headers)
    style_header_row(ws)
    
    # Data
    for invoice in invoices_queryset:
        ws.append([
            invoice.invoice_number,
            invoice.party.name,
            invoice.get_invoice_type_display(),
            str(invoice.issue_date) if invoice.issue_date else '',
            str(invoice.due_date) if invoice.due_date else '',
            invoice.get_status_display(),
            str(invoice.get_total()),
            str(invoice.get_paid_amount()),
            str(invoice.get_remaining())
        ])
    
    auto_adjust_columns(ws)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_payments_excel(payments_queryset):
    """Create Excel file for payments"""
    check_excel_dependencies()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "الدفعات"
    ws.sheet_view.rightToLeft = True
    
    # Headers
    headers = ['المعرف', 'رقم الفاتورة', 'الطرف', 'المبلغ', 'طريقة الدفع', 
               'تاريخ الدفع', 'ملاحظات']
    ws.append(headers)
    style_header_row(ws)
    
    # Data
    for payment in payments_queryset:
        ws.append([
            payment.id,
            payment.invoice.invoice_number,
            payment.invoice.party.name,
            str(payment.amount),
            payment.get_payment_method_display(),
            str(payment.payment_date) if payment.payment_date else '',
            payment.notes or ''
        ])
    
    auto_adjust_columns(ws)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_ledger_excel(ledger_queryset):
    """Create Excel file for ledger entries"""
    check_excel_dependencies()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "دفتر الأستاذ"
    ws.sheet_view.rightToLeft = True
    
    # Headers
    headers = ['المعرف', 'الطرف', 'نوع القيد', 'المبلغ', 'التاريخ', 
               'رقم الفاتورة', 'الوصف']
    ws.append(headers)
    style_header_row(ws)
    
    # Data
    for entry in ledger_queryset:
        ws.append([
            entry.id,
            entry.party.name,
            entry.get_entry_type_display(),
            str(entry.amount),
            str(entry.date) if entry.date else '',
            entry.invoice.invoice_number if entry.invoice else '',
            entry.description or ''
        ])
    
    auto_adjust_columns(ws)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_full_export_excel(parties, invoices, payments, ledger_entries):
    """Create Excel file with all data in separate sheets"""
    check_excel_dependencies()
    
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Parties sheet
    ws_parties = wb.create_sheet("الأطراف")
    ws_parties.sheet_view.rightToLeft = True
    ws_parties.append(['المعرف', 'الاسم', 'النوع', 'الهاتف', 'البريد الإلكتروني', 'الرصيد'])
    style_header_row(ws_parties)
    for party in parties:
        ws_parties.append([
            party.id, party.name, party.get_party_type_display(),
            party.phone or '', party.email or '', str(party.get_balance())
        ])
    auto_adjust_columns(ws_parties)
    
    # Invoices sheet
    ws_invoices = wb.create_sheet("الفواتير")
    ws_invoices.sheet_view.rightToLeft = True
    ws_invoices.append(['رقم الفاتورة', 'الطرف', 'النوع', 'تاريخ الإصدار', 'الحالة', 'المجموع', 'المتبقي'])
    style_header_row(ws_invoices)
    for invoice in invoices:
        ws_invoices.append([
            invoice.invoice_number, invoice.party.name, invoice.get_invoice_type_display(),
            str(invoice.issue_date), invoice.get_status_display(),
            str(invoice.get_total()), str(invoice.get_remaining())
        ])
    auto_adjust_columns(ws_invoices)
    
    # Payments sheet
    ws_payments = wb.create_sheet("الدفعات")
    ws_payments.sheet_view.rightToLeft = True
    ws_payments.append(['رقم الفاتورة', 'الطرف', 'المبلغ', 'طريقة الدفع', 'تاريخ الدفع'])
    style_header_row(ws_payments)
    for payment in payments:
        ws_payments.append([
            payment.invoice.invoice_number, payment.invoice.party.name,
            str(payment.amount), payment.get_payment_method_display(),
            str(payment.payment_date)
        ])
    auto_adjust_columns(ws_payments)
    
    # Ledger sheet
    ws_ledger = wb.create_sheet("دفتر الأستاذ")
    ws_ledger.sheet_view.rightToLeft = True
    ws_ledger.append(['الطرف', 'نوع القيد', 'المبلغ', 'التاريخ', 'الوصف'])
    style_header_row(ws_ledger)
    for entry in ledger_entries:
        ws_ledger.append([
            entry.party.name, entry.get_entry_type_display(),
            str(entry.amount), str(entry.date), entry.description or ''
        ])
    auto_adjust_columns(ws_ledger)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_import_template(resource_type):
    """Create import template Excel file"""
    check_excel_dependencies()
    
    wb = Workbook()
    ws = wb.active
    ws.sheet_view.rightToLeft = True
    
    if resource_type == 'parties':
        ws.title = "قالب الأطراف"
        headers = ['الاسم', 'النوع (customer/supplier)', 'الهاتف', 'البريد الإلكتروني']
        example = ['شركة الاختبار', 'customer', '0501234567', 'test@example.com']
    elif resource_type == 'invoices':
        ws.title = "قالب الفواتير"
        headers = ['معرف الطرف', 'النوع (sale/purchase)', 'تاريخ الإصدار', 'تاريخ الاستحقاق', 'ملاحظات']
        example = ['1', 'sale', '2024-01-15', '2024-02-15', 'فاتورة اختبار']
    elif resource_type == 'payments':
        ws.title = "قالب الدفعات"
        headers = ['معرف الفاتورة', 'المبلغ', 'طريقة الدفع (cash/card/transfer)', 'تاريخ الدفع', 'ملاحظات']
        example = ['1', '1000.00', 'cash', '2024-01-15', 'دفعة اختبار']
    else:
        ws.title = "قالب"
        headers = ['حقل 1', 'حقل 2']
        example = ['قيمة 1', 'قيمة 2']
    
    ws.append(headers)
    style_header_row(ws)
    ws.append(example)
    auto_adjust_columns(ws)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def parse_parties_excel(file):
    """Parse parties from Excel file"""
    check_excel_dependencies()
    
    df = pd.read_excel(file, engine='openpyxl')
    
    # Map column names (support both Arabic and English)
    column_map = {
        'الاسم': 'name', 'name': 'name',
        'النوع (customer/supplier)': 'party_type', 'النوع': 'party_type', 
        'party_type': 'party_type', 'type': 'party_type',
        'الهاتف': 'phone', 'phone': 'phone',
        'البريد الإلكتروني': 'email', 'email': 'email',
    }
    
    df = df.rename(columns=column_map)
    
    parties_data = []
    errors = []
    
    for idx, row in df.iterrows():
        try:
            party_data = {
                'name': str(row.get('name', '')).strip(),
                'party_type': str(row.get('party_type', '')).strip().lower(),
                'phone': str(row.get('phone', '')).strip() if pd.notna(row.get('phone')) else None,
                'email': str(row.get('email', '')).strip() if pd.notna(row.get('email')) else None,
            }
            
            if not party_data['name']:
                errors.append(f"Row {idx + 2}: Name is required")
                continue
            
            if party_data['party_type'] not in ['customer', 'supplier']:
                errors.append(f"Row {idx + 2}: Invalid party type '{party_data['party_type']}'")
                continue
            
            parties_data.append(party_data)
        except Exception as e:
            errors.append(f"Row {idx + 2}: {str(e)}")
    
    return parties_data, errors
